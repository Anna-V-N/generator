from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta

#generator for reading big file
def xlsx_line_reader(file_name):
    workbook = load_workbook(filename=file_name)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):  #skip headings
        yield row

#main filter function
def filter_yesterday_xlsx(input_file, output_file):
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")

    #creating new Excel file
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ID", "Date", "Value"])  #headings

    #reading file and filtering
    for row in xlsx_line_reader(input_file):
        record_date = row[1]
        if record_date == yesterday:
            sheet.append(row)

    #saving filtered file
    workbook.save(output_file)

#call the function and filtering
filter_yesterday_xlsx("large_file.xlsx", "filtered_file.xlsx")
